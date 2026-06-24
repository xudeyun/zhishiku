"""Core knowledge base engine with Supabase PostgreSQL + pgvector for persistent storage."""

import os
import uuid
import json
from typing import Optional

import psycopg2
from psycopg2.extras import RealDictCursor
from sentence_transformers import SentenceTransformer

# ── Config (from Streamlit secrets or env) ──

DATABASE_URL = os.environ.get("DATABASE_URL", "")
EMBEDDING_MODEL_NAME = "paraphrase-multilingual-MiniLM-L12-v2"

MODULES = [
    {"id": "accounting", "name": "核算", "icon": "📊", "description": "会计核算相关知识"},
    {"id": "tax", "name": "税务", "icon": "🧾", "description": "税务法规与实务知识"},
    {"id": "other", "name": "其他", "icon": "📁", "description": "其他综合知识"},
]

CHUNK_SIZE = 500
CHUNK_OVERLAP = 100

# ── Database connection ──

_conn = None

def _get_conn():
    global _conn
    if _conn is None or _conn.closed:
        _conn = psycopg2.connect(DATABASE_URL, cursor_factory=RealDictCursor)
    return _conn


# ── Embedding model (lazy load) ──

_embed_model = None

def _get_embed_model():
    global _embed_model
    if _embed_model is None:
        _embed_model = SentenceTransformer(EMBEDING_MODEL_NAME)
    return _embed_model


def embed_text(text: str) -> list[float]:
    model = _get_embed_model()
    return model.encode(text).tolist()


def embed_batch(texts: list[str]) -> list[list[float]]:
    model = _get_embed_model()
    return model.encode(texts).tolist()


# ── Document parsing ──

def parse_pdf(file_bytes: bytes) -> list[dict]:
    import pdfplumber
    import io
    elements = []
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for i, page in enumerate(pdf.pages):
            text = page.extract_text() or ""
            if text.strip():
                elements.append({"content": text, "page_number": i + 1})
    return elements


def parse_docx(file_bytes: bytes) -> list[dict]:
    from docx import Document
    import io
    doc = Document(io.BytesIO(file_bytes))
    elements = []
    section = None
    for para in doc.paragraphs:
        if para.style and para.style.name.startswith("Heading"):
            section = para.text.strip()
            continue
        if para.text.strip():
            elements.append({"content": para.text.strip(), "section_title": section})
    return elements


def parse_xlsx(file_bytes: bytes) -> list[dict]:
    from openpyxl import load_workbook
    import io
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    elements = []
    for sheet in wb.worksheets:
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
            cells = [str(c) if c is not None else "" for c in row]
            line = " | ".join(cells)
            if line.strip(" |"):
                elements.append({"content": line, "sheet_name": sheet.title, "row_number": row_idx})
    wb.close()
    return elements


def parse_txt(file_bytes: bytes) -> list[dict]:
    import chardet
    enc = (chardet.detect(file_bytes).get("encoding") or "utf-8")
    try:
        text = file_bytes.decode(enc)
    except (UnicodeDecodeError, LookupError):
        text = file_bytes.decode("utf-8", errors="replace")
    return [{"content": text}]


def parse_md(file_bytes: bytes) -> list[dict]:
    text = file_bytes.decode("utf-8", errors="replace")
    elements = []
    section = None
    for line in text.split("\n"):
        if line.startswith("#"):
            section = line.lstrip("#").strip()
            continue
        if line.strip():
            elements.append({"content": line, "section_title": section})
    return elements


PARSERS = {"pdf": parse_pdf, "docx": parse_docx, "xlsx": parse_xlsx, "txt": parse_txt, "md": parse_md}


# ── Chunking ──

def chunk_text(text: str, size=CHUNK_SIZE, overlap=CHUNK_OVERLAP) -> list[str]:
    if len(text) <= size:
        return [text] if text.strip() else []
    chunks = []
    start = 0
    while start < len(text):
        end = min(start + size, len(text))
        chunks.append(text[start:end].strip())
        start = end - overlap
        if start >= len(text):
            break
    return [c for c in chunks if c]


def chunk_elements(elements: list[dict]) -> list[dict]:
    all_chunks = []
    idx = 0
    for elem in elements:
        text = elem.get("content", "")
        if not text.strip():
            continue
        splits = chunk_text(text)
        for s in splits:
            all_chunks.append({
                "content": s,
                "chunk_index": idx,
                "page_number": elem.get("page_number"),
                "section_title": elem.get("section_title"),
                "sheet_name": elem.get("sheet_name"),
                "row_number": elem.get("row_number"),
            })
            idx += 1
    return all_chunks


# ── Indexing ──

def _vec_to_sql(vec: list[float]) -> str:
    return "[" + ",".join(str(v) for v in vec) + "]"


def index_chunks(chunks: list[dict], source_type: str, source_id: str, module_id: str):
    if not chunks:
        return
    texts = [c["content"] for c in chunks]
    embeddings = embed_batch(texts)

    conn = _get_conn()
    for i, chunk in enumerate(chunks):
        emb = embeddings[i] if i < len(embeddings) else None
        conn.execute(
            """INSERT INTO chunks (id, source_type, source_id, module_id, content, chunk_index,
               page_number, section_title, sheet_name, row_number, embedding)
               VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
            (uuid.uuid4().hex, source_type, source_id, module_id,
             chunk["content"], chunk["chunk_index"],
             chunk.get("page_number"), chunk.get("section_title"),
             chunk.get("sheet_name"), chunk.get("row_number"),
             _vec_to_sql(emb) if emb else None),
        )
    conn.commit()


def delete_source_chunks(source_type: str, source_id: str):
    conn = _get_conn()
    conn.execute("DELETE FROM chunks WHERE source_type=%s AND source_id=%s", (source_type, source_id))
    conn.commit()


# ── Retrieval (pgvector) ──

def search(query: str, module_id: str | None = None, top_k: int = 8) -> list[dict]:
    query_vec = embed_text(query)
    vec_str = _vec_to_sql(query_vec)

    conn = _get_conn()
    if module_id:
        conn.execute(
            """SELECT c.*, d.filename AS doc_filename, k.title AS kn_title,
                      1 - (c.embedding <=> %s::vector) AS score
               FROM chunks c
               LEFT JOIN documents d ON c.source_type='document' AND c.source_id = d.id
               LEFT JOIN knowledge k ON c.source_type='knowledge' AND c.source_id = k.id
               WHERE c.module_id = %s AND c.embedding IS NOT NULL
               ORDER BY c.embedding <=> %s::vector
               LIMIT %s""",
            (vec_str, module_id, vec_str, top_k),
        )
    else:
        conn.execute(
            """SELECT c.*, d.filename AS doc_filename, k.title AS kn_title,
                      1 - (c.embedding <=> %s::vector) AS score
               FROM chunks c
               LEFT JOIN documents d ON c.source_type='document' AND c.source_id = d.id
               LEFT JOIN knowledge k ON c.source_type='knowledge' AND c.source_id = k.id
               WHERE c.embedding IS NOT NULL
               ORDER BY c.embedding <=> %s::vector
               LIMIT %s""",
            (vec_str, vec_str, top_k),
        )
    rows = conn.fetchall()
    result = []
    for r in rows:
        source_title = r.get("doc_filename") or r.get("kn_title") or "未知"
        result.append({
            "chunk_id": r["id"],
            "content": r["content"],
            "score": float(r.get("score", 0)),
            "source_type": r["source_type"],
            "source_id": r["source_id"],
            "module_id": r["module_id"],
            "page_number": r.get("page_number"),
            "section_title": r.get("section_title"),
            "sheet_name": r.get("sheet_name"),
            "row_number": r.get("row_number"),
            "source_title": source_title,
        })
    return result


# ── Document operations ──

def add_document(module_id: str, filename: str, file_type: str, file_bytes: bytes, title: str = "") -> str:
    doc_id = uuid.uuid4().hex
    conn = _get_conn()
    conn.execute(
        "INSERT INTO documents (id, module_id, filename, file_type, title, status) VALUES (%s,%s,%s,%s,%s,'processing')",
        (doc_id, module_id, filename, file_type, title or filename.rsplit(".", 1)[0]),
    )
    conn.commit()

    try:
        parser = PARSERS.get(file_type)
        if not parser:
            raise ValueError(f"不支持的文件类型: {file_type}")
        elements = parser(file_bytes)
        chunks = chunk_elements(elements)
        index_chunks(chunks, "document", doc_id, module_id)

        full_text = "\n\n".join(e.get("content", "") for e in elements)
        conn.execute("UPDATE documents SET status='ready', file_content=%s WHERE id=%s", (full_text, doc_id))
        conn.commit()
    except Exception as e:
        conn.execute("UPDATE documents SET status='error' WHERE id=%s", (doc_id,))
        conn.commit()
        raise

    return doc_id


def list_documents(module_id: str) -> list[dict]:
    conn = _get_conn()
    conn.execute("SELECT id, module_id, filename, file_type, title, status, created_at FROM documents WHERE module_id=%s ORDER BY created_at DESC", (module_id,))
    return [dict(r) for r in conn.fetchall()]


def delete_document(doc_id: str):
    conn = _get_conn()
    conn.execute("DELETE FROM documents WHERE id=%s", (doc_id,))
    conn.commit()
    delete_source_chunks("document", doc_id)


# ── Knowledge operations ──

def add_knowledge(module_id: str, title: str, content: str, tags: str = "") -> str:
    kn_id = uuid.uuid4().hex
    conn = _get_conn()
    conn.execute(
        "INSERT INTO knowledge (id, module_id, title, content, tags) VALUES (%s,%s,%s,%s,%s)",
        (kn_id, module_id, title, content, tags),
    )
    conn.commit()

    elements = [{"content": content}]
    chunks = chunk_elements(elements)
    index_chunks(chunks, "knowledge", kn_id, module_id)
    return kn_id


def list_knowledge(module_id: str) -> list[dict]:
    conn = _get_conn()
    conn.execute("SELECT * FROM knowledge WHERE module_id=%s ORDER BY updated_at DESC", (module_id,))
    return [dict(r) for r in conn.fetchall()]


def update_knowledge(kn_id: str, title: str = None, content: str = None, tags: str = None):
    conn = _get_conn()
    kn = conn.execute("SELECT * FROM knowledge WHERE id=%s", (kn_id,)).fetchone()
    if not kn:
        return
    t = title if title is not None else kn["title"]
    c = content if content is not None else kn["content"]
    tg = tags if tags is not None else kn["tags"]
    conn.execute(
        "UPDATE knowledge SET title=%s, content=%s, tags=%s, version=version+1, updated_at=now() WHERE id=%s",
        (t, c, tg, kn_id),
    )
    conn.commit()

    delete_source_chunks("knowledge", kn_id)
    elements = [{"content": c}]
    chunks = chunk_elements(elements)
    index_chunks(chunks, "knowledge", kn_id, kn["module_id"])


def delete_knowledge(kn_id: str):
    conn = _get_conn()
    conn.execute("DELETE FROM knowledge WHERE id=%s", (kn_id,))
    conn.commit()
    delete_source_chunks("knowledge", kn_id)


def get_knowledge(kn_id: str) -> dict | None:
    conn = _get_conn()
    conn.execute("SELECT * FROM knowledge WHERE id=%s", (kn_id,))
    r = conn.fetchone()
    return dict(r) if r else None


# ── Conversation operations ──

def create_conversation(module_id: str, title: str = "") -> str:
    conv_id = uuid.uuid4().hex
    conn = _get_conn()
    conn.execute(
        "INSERT INTO conversations (id, module_id, title) VALUES (%s,%s,%s)",
        (conv_id, module_id, title or "新对话"),
    )
    conn.commit()
    return conv_id


def add_message(conversation_id: str, role: str, content: str, citations: list = None):
    msg_id = uuid.uuid4().hex
    conn = _get_conn()
    conn.execute(
        "INSERT INTO messages (id, conversation_id, role, content, citations) VALUES (%s,%s,%s,%s,%s)",
        (msg_id, conversation_id, role, content, json.dumps(citations or [], ensure_ascii=False)),
    )
    conn.execute("UPDATE conversations SET updated_at=now() WHERE id=%s", (conversation_id,))
    conn.commit()


def get_conversation_messages(conv_id: str) -> list[dict]:
    conn = _get_conn()
    conn.execute("SELECT * FROM messages WHERE conversation_id=%s ORDER BY created_at", (conv_id,))
    return [dict(r) for r in conn.fetchall()]


def list_conversations(module_id: str) -> list[dict]:
    conn = _get_conn()
    conn.execute("SELECT * FROM conversations WHERE module_id=%s ORDER BY updated_at DESC", (module_id,))
    return [dict(r) for r in conn.fetchall()]


def delete_conversation(conv_id: str):
    conn = _get_conn()
    conn.execute("DELETE FROM messages WHERE conversation_id=%s", (conv_id,))
    conn.execute("DELETE FROM conversations WHERE id=%s", (conv_id,))
    conn.commit()
